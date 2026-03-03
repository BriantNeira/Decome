"""Tests for bulk reminder import service and multi-entity imports."""
import io
import uuid
import datetime
import pytest
from unittest.mock import AsyncMock, MagicMock, patch

from app.services.import_service import (
    _parse_date,
    generate_template_xlsx,
    generate_entity_template_xlsx,
    parse_and_validate,
    parse_and_validate_entities,
    create_entities_from_rows,
)


# --- Unit tests for _parse_date ---

def test_parse_date_iso():
    assert _parse_date("2026-06-01") == datetime.date(2026, 6, 1)


def test_parse_date_us_format():
    assert _parse_date("6/1/2026") == datetime.date(2026, 6, 1)


def test_parse_date_eu_format():
    assert _parse_date("1/6/2026") == datetime.date(2026, 1, 6)


def test_parse_date_invalid():
    assert _parse_date("not-a-date") is None


def test_parse_date_empty():
    assert _parse_date("") is None


# --- Unit test for generate_template_xlsx ---

def test_generate_template_xlsx():
    xlsx_bytes = generate_template_xlsx()
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 1000  # meaningful content
    # Should be a valid xlsx (starts with PK magic bytes)
    assert xlsx_bytes[:2] == b"PK"


# --- Tests for parse_and_validate ---

def _make_xlsx_bytes(rows: list[list]) -> bytes:
    """Helper: create a minimal xlsx with given rows (first row = headers)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_mock_db(accounts=None, programs=None, types=None):
    """Build a mock AsyncSession returning the given lookup data."""
    mock_db = AsyncMock()
    call_count = [0]

    # Default empty lists
    accs = accounts or []
    progs = programs or []
    typs = types or []

    async def mock_execute(stmt, *args, **kwargs):
        result = MagicMock()
        n = call_count[0]
        call_count[0] += 1
        if n == 0:
            result.scalars.return_value.all.return_value = accs
        elif n == 1:
            result.scalars.return_value.all.return_value = progs
        elif n == 2:
            result.scalars.return_value.all.return_value = typs
        else:
            result.scalars.return_value.all.return_value = []
        return result

    mock_db.execute = mock_execute
    return mock_db


def _make_account(name: str, account_id: uuid.UUID = None) -> MagicMock:
    a = MagicMock()
    a.id = account_id or uuid.uuid4()
    a.name = name
    a.is_active = True
    return a


def _make_program(name: str, account_id: uuid.UUID, program_id: uuid.UUID = None) -> MagicMock:
    p = MagicMock()
    p.id = program_id or uuid.uuid4()
    p.name = name
    p.account_id = account_id
    p.is_active = True
    return p


def _make_type(name: str, type_id: int = 1) -> MagicMock:
    t = MagicMock()
    t.id = type_id
    t.name = name
    t.is_active = True
    return t


@pytest.mark.asyncio
async def test_parse_and_validate_valid_row():
    """A valid row returns status=ok with resolved FKs."""
    acc = _make_account("Acme Corp")
    prog = _make_program("Q1 Promo", acc.id)
    rt = _make_type("Seasonal Order", 3)

    headers = ["account", "program", "reminder_type", "title", "due_date", "notes", "recurrence"]
    data_row = ["Acme Corp", "Q1 Promo", "Seasonal Order", "Spring Order", "2026-04-01", "Notes here", "none"]
    xlsx = _make_xlsx_bytes([headers, data_row])

    db = _make_mock_db(accounts=[acc], programs=[prog], types=[rt])
    rows = await parse_and_validate(db, xlsx)

    assert len(rows) == 1
    assert rows[0].status == "ok"
    assert rows[0].account_id == acc.id
    assert rows[0].program_id == prog.id
    assert rows[0].type_id == 3
    assert rows[0].due_date == datetime.date(2026, 4, 1)
    assert rows[0].title == "Spring Order"


@pytest.mark.asyncio
async def test_parse_and_validate_unknown_account():
    """A row with an unknown account name returns status=error."""
    headers = ["account", "program", "reminder_type", "title", "due_date", "notes", "recurrence"]
    data_row = ["Unknown Inc", "Q1 Promo", "Seasonal Order", "Spring Order", "2026-04-01", "", "none"]
    xlsx = _make_xlsx_bytes([headers, data_row])

    db = _make_mock_db(accounts=[], programs=[], types=[])
    rows = await parse_and_validate(db, xlsx)

    assert len(rows) == 1
    assert rows[0].status == "error"
    assert "Account" in rows[0].error_msg or "account" in rows[0].error_msg.lower()


@pytest.mark.asyncio
async def test_parse_and_validate_invalid_date():
    """A row with invalid date returns status=error."""
    acc = _make_account("Acme Corp")
    prog = _make_program("Q1 Promo", acc.id)
    rt = _make_type("Seasonal Order", 3)

    headers = ["account", "program", "reminder_type", "title", "due_date", "notes", "recurrence"]
    data_row = ["Acme Corp", "Q1 Promo", "Seasonal Order", "Spring Order", "not-a-date", "", "none"]
    xlsx = _make_xlsx_bytes([headers, data_row])

    db = _make_mock_db(accounts=[acc], programs=[prog], types=[rt])
    rows = await parse_and_validate(db, xlsx)

    assert len(rows) == 1
    assert rows[0].status == "error"
    assert "due_date" in rows[0].error_msg.lower() or "date" in rows[0].error_msg.lower()


@pytest.mark.asyncio
async def test_parse_and_validate_missing_required_field():
    """A row missing required 'title' returns status=error."""
    acc = _make_account("Acme Corp")
    prog = _make_program("Q1 Promo", acc.id)
    rt = _make_type("Seasonal Order", 3)

    headers = ["account", "program", "reminder_type", "title", "due_date", "notes", "recurrence"]
    data_row = ["Acme Corp", "Q1 Promo", "Seasonal Order", "", "2026-04-01", "", "none"]  # empty title
    xlsx = _make_xlsx_bytes([headers, data_row])

    db = _make_mock_db(accounts=[acc], programs=[prog], types=[rt])
    rows = await parse_and_validate(db, xlsx)

    assert len(rows) == 1
    assert rows[0].status == "error"


@pytest.mark.asyncio
async def test_parse_and_validate_multiple_rows():
    """Multiple rows each validated independently."""
    acc = _make_account("Acme Corp")
    prog = _make_program("Q1 Promo", acc.id)
    rt = _make_type("Seasonal Order", 3)

    headers = ["account", "program", "reminder_type", "title", "due_date", "notes", "recurrence"]
    rows_data = [
        ["Acme Corp", "Q1 Promo", "Seasonal Order", "Order 1", "2026-04-01", "", "none"],
        ["Acme Corp", "Q1 Promo", "Seasonal Order", "Order 2", "2026-05-01", "", "monthly"],
        ["Unknown Inc", "Q1 Promo", "Seasonal Order", "Order 3", "2026-06-01", "", "none"],  # bad account
    ]
    xlsx = _make_xlsx_bytes([headers] + rows_data)

    db = _make_mock_db(accounts=[acc], programs=[prog], types=[rt])
    rows = await parse_and_validate(db, xlsx)

    assert len(rows) == 3
    assert rows[0].status == "ok"
    assert rows[1].status == "ok"
    assert rows[2].status == "error"


@pytest.mark.asyncio
async def test_parse_and_validate_skips_blank_rows():
    """Completely blank rows are skipped."""
    acc = _make_account("Acme Corp")
    prog = _make_program("Q1 Promo", acc.id)
    rt = _make_type("Seasonal Order", 3)

    headers = ["account", "program", "reminder_type", "title", "due_date", "notes", "recurrence"]
    xlsx = _make_xlsx_bytes([
        headers,
        ["Acme Corp", "Q1 Promo", "Seasonal Order", "Order 1", "2026-04-01", "", "none"],
        [None, None, None, None, None, None, None],  # blank row
        ["Acme Corp", "Q1 Promo", "Seasonal Order", "Order 2", "2026-05-01", "", "none"],
    ])

    db = _make_mock_db(accounts=[acc], programs=[prog], types=[rt])
    rows = await parse_and_validate(db, xlsx)

    assert len(rows) == 2  # blank row skipped


def test_generate_template_xlsx_has_correct_headers():
    """Template xlsx should have the correct column headers."""
    import openpyxl
    xlsx_bytes = generate_template_xlsx()
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    # All required fields should be present in header labels
    header_str = " ".join(str(h).lower() for h in headers if h)
    assert "account" in header_str
    assert "program" in header_str
    assert "title" in header_str
    assert "due" in header_str


# ---------------------------------------------------------------------------
# Entity template tests
# ---------------------------------------------------------------------------

def test_generate_entity_template_accounts():
    """Template for accounts should be a valid xlsx."""
    xlsx = generate_entity_template_xlsx("accounts")
    assert isinstance(xlsx, bytes)
    assert xlsx[:2] == b"PK"


def test_generate_entity_template_programs():
    xlsx = generate_entity_template_xlsx("programs")
    assert isinstance(xlsx, bytes) and len(xlsx) > 500


def test_generate_entity_template_contacts():
    xlsx = generate_entity_template_xlsx("contacts")
    assert isinstance(xlsx, bytes) and len(xlsx) > 500


def test_generate_entity_template_assignments():
    xlsx = generate_entity_template_xlsx("assignments")
    assert isinstance(xlsx, bytes) and len(xlsx) > 500


def test_generate_entity_template_unknown_raises():
    with pytest.raises(ValueError, match="Unknown entity type"):
        generate_entity_template_xlsx("widgets")


# ---------------------------------------------------------------------------
# Entity-specific mock DB helpers
# ---------------------------------------------------------------------------

def _make_user(email: str, user_id: uuid.UUID = None) -> MagicMock:
    u = MagicMock()
    u.id = user_id or uuid.uuid4()
    u.email = email
    u.full_name = email.split("@")[0].title()
    u.is_active = True
    return u


def _make_entity_mock_db(
    accounts=None,
    programs=None,
    users=None,
    assignments=None,
):
    """Build a mock AsyncSession that returns lookups in sequence for entity validators.

    Entity validators call db.execute multiple times depending on entity type.
    We return objects based on call sequence.
    """
    mock_db = AsyncMock()
    # We'll gather all call results and return them in order
    result_sets = []

    # accounts import: 1 call (existing accounts)
    # programs import: 2 calls (existing programs, accounts)
    # contacts import: 1 call (accounts)
    # assignments import: 4 calls (users, accounts, programs, existing assignments)

    # To be flexible, we store a list of result sets and pop from front
    if accounts is not None:
        result_sets.append(accounts)
    if programs is not None:
        result_sets.append(programs)
    if users is not None:
        result_sets.append(users)
    if assignments is not None:
        result_sets.append(assignments)

    call_idx = [0]

    async def mock_execute(stmt, *args, **kwargs):
        result = MagicMock()
        n = call_idx[0]
        call_idx[0] += 1
        if n < len(result_sets):
            result.scalars.return_value.all.return_value = result_sets[n]
        else:
            result.scalars.return_value.all.return_value = []
        return result

    mock_db.execute = mock_execute
    return mock_db


# ---------------------------------------------------------------------------
# Account import tests
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_import_accounts_dry_run():
    """Validate accounts: new name = ok, existing = skipped."""
    existing_acct = _make_account("Acme Corp")

    # Account validator calls db.execute once: select(Account)
    db = AsyncMock()
    call_idx = [0]

    async def mock_execute(stmt, *args, **kwargs):
        n = call_idx[0]
        call_idx[0] += 1
        result = MagicMock()
        if n == 0:
            result.scalars.return_value.all.return_value = [existing_acct]
        else:
            result.scalars.return_value.all.return_value = []
        return result

    db.execute = mock_execute

    headers = ["Name *", "Code", "Description"]
    rows_data = [
        ["Acme Corp", "ACME-01", "Existing account"],      # should be skipped
        ["NewCo", "NEW-01", "Brand new account"],           # should be ok
    ]
    xlsx = _make_xlsx_bytes([headers] + rows_data)

    rows = await parse_and_validate_entities(db, xlsx, "accounts")

    assert len(rows) == 2
    assert rows[0].status == "skipped"
    assert "already exists" in rows[0].error_msg
    assert rows[1].status == "ok"


@pytest.mark.asyncio
async def test_import_accounts_missing_name():
    """Account without a name should be an error."""
    db = AsyncMock()

    async def mock_execute(stmt, *args, **kwargs):
        result = MagicMock()
        result.scalars.return_value.all.return_value = []
        return result

    db.execute = mock_execute

    headers = ["Name *", "Code", "Description"]
    rows_data = [
        ["", "CODE1", "Missing name"],
    ]
    xlsx = _make_xlsx_bytes([headers] + rows_data)

    rows = await parse_and_validate_entities(db, xlsx, "accounts")
    assert len(rows) == 1
    assert rows[0].status == "error"
    assert "name" in rows[0].error_msg.lower()


# ---------------------------------------------------------------------------
# Program import tests
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_import_programs_with_season():
    """Programs should parse and store season value."""
    existing_acct = _make_account("Acme Corp")

    db = AsyncMock()
    call_idx = [0]

    async def mock_execute(stmt, *args, **kwargs):
        n = call_idx[0]
        call_idx[0] += 1
        result = MagicMock()
        if n == 0:
            # existing programs
            result.scalars.return_value.all.return_value = []
        elif n == 1:
            # accounts for resolution
            result.scalars.return_value.all.return_value = [existing_acct]
        else:
            result.scalars.return_value.all.return_value = []
        return result

    db.execute = mock_execute

    headers = ["Name *", "Description", "Account", "Season"]
    rows_data = [
        ["Spring Campaign", "Q1 promo", "Acme Corp", "Spring 2026"],
        ["Fall Review", "Annual review", "Acme Corp", "Fall 2026"],
    ]
    xlsx = _make_xlsx_bytes([headers] + rows_data)

    rows = await parse_and_validate_entities(db, xlsx, "programs")
    assert len(rows) == 2
    assert rows[0].status == "ok"
    assert rows[0].recurrence == "Spring 2026"  # season stored in recurrence field
    assert rows[1].status == "ok"
    assert rows[1].recurrence == "Fall 2026"


@pytest.mark.asyncio
async def test_import_programs_duplicate_skipped():
    """Duplicate program names should be skipped."""
    existing_prog = MagicMock()
    existing_prog.name = "Existing Program"
    existing_prog.id = uuid.uuid4()

    db = AsyncMock()
    call_idx = [0]

    async def mock_execute(stmt, *args, **kwargs):
        n = call_idx[0]
        call_idx[0] += 1
        result = MagicMock()
        if n == 0:
            result.scalars.return_value.all.return_value = [existing_prog]
        elif n == 1:
            result.scalars.return_value.all.return_value = []  # accounts
        else:
            result.scalars.return_value.all.return_value = []
        return result

    db.execute = mock_execute

    headers = ["Name *", "Description", "Account", "Season"]
    rows_data = [
        ["Existing Program", "Already there", "", ""],
        ["New Program", "Fresh one", "", "Summer"],
    ]
    xlsx = _make_xlsx_bytes([headers] + rows_data)

    rows = await parse_and_validate_entities(db, xlsx, "programs")
    assert len(rows) == 2
    assert rows[0].status == "skipped"
    assert rows[1].status == "ok"


# ---------------------------------------------------------------------------
# Contact import tests
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_import_contacts_resolves_account():
    """Contact import should resolve account by name and validate required fields."""
    acct = _make_account("Acme Corp")

    db = AsyncMock()
    call_idx = [0]

    async def mock_execute(stmt, *args, **kwargs):
        n = call_idx[0]
        call_idx[0] += 1
        result = MagicMock()
        if n == 0:
            result.scalars.return_value.all.return_value = [acct]
        else:
            result.scalars.return_value.all.return_value = []
        return result

    db.execute = mock_execute

    headers = ["Account *", "First Name *", "Last Name *", "Email", "Phone", "Title", "Decision Maker"]
    rows_data = [
        ["Acme Corp", "Jane", "Doe", "jane@acme.com", "555-1234", "Buyer", "true"],
        ["Unknown Co", "John", "Smith", "", "", "", "false"],  # unknown account
    ]
    xlsx = _make_xlsx_bytes([headers] + rows_data)

    rows = await parse_and_validate_entities(db, xlsx, "contacts")
    assert len(rows) == 2
    assert rows[0].status == "ok"
    assert rows[0].account_id == acct.id
    assert rows[0].title == "Jane Doe"  # name stored in title
    assert rows[1].status == "error"
    assert "Account" in rows[1].error_msg


@pytest.mark.asyncio
async def test_import_contacts_missing_required():
    """Contacts missing required fields should have errors."""
    acct = _make_account("Acme Corp")

    db = AsyncMock()

    async def mock_execute(stmt, *args, **kwargs):
        result = MagicMock()
        result.scalars.return_value.all.return_value = [acct]
        return result

    db.execute = mock_execute

    headers = ["Account *", "First Name *", "Last Name *", "Email", "Phone", "Title", "Decision Maker"]
    rows_data = [
        ["Acme Corp", "", "Doe", "", "", "", ""],  # missing first_name
    ]
    xlsx = _make_xlsx_bytes([headers] + rows_data)

    rows = await parse_and_validate_entities(db, xlsx, "contacts")
    assert len(rows) == 1
    assert rows[0].status == "error"
    assert "first_name" in rows[0].error_msg.lower()


# ---------------------------------------------------------------------------
# Assignment import tests
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_import_assignments_skips_duplicates():
    """Existing assignments should be skipped."""
    user = _make_user("bdm@example.com")
    acct = _make_account("Acme Corp")
    prog = _make_program("Q1 Promo", acct.id)

    existing_assignment = MagicMock()
    existing_assignment.user_id = user.id
    existing_assignment.account_id = acct.id
    existing_assignment.program_id = prog.id
    existing_assignment.is_active = True

    db = AsyncMock()
    call_idx = [0]

    async def mock_execute(stmt, *args, **kwargs):
        n = call_idx[0]
        call_idx[0] += 1
        result = MagicMock()
        if n == 0:
            result.scalars.return_value.all.return_value = [user]  # users
        elif n == 1:
            result.scalars.return_value.all.return_value = [acct]  # accounts
        elif n == 2:
            result.scalars.return_value.all.return_value = [prog]  # programs
        elif n == 3:
            result.scalars.return_value.all.return_value = [existing_assignment]  # existing assignments
        else:
            result.scalars.return_value.all.return_value = []
        return result

    db.execute = mock_execute

    headers = ["BDM Email *", "Account *", "Program *"]
    rows_data = [
        ["bdm@example.com", "Acme Corp", "Q1 Promo"],  # duplicate - should be skipped
    ]
    xlsx = _make_xlsx_bytes([headers] + rows_data)

    rows = await parse_and_validate_entities(db, xlsx, "assignments")
    assert len(rows) == 1
    assert rows[0].status == "skipped"
    assert "already exists" in rows[0].error_msg.lower()


@pytest.mark.asyncio
async def test_import_assignments_valid():
    """New assignment with valid BDM, account, and program should be ok."""
    user = _make_user("bdm@example.com")
    acct = _make_account("Acme Corp")
    prog = _make_program("Q1 Promo", acct.id)

    db = AsyncMock()
    call_idx = [0]

    async def mock_execute(stmt, *args, **kwargs):
        n = call_idx[0]
        call_idx[0] += 1
        result = MagicMock()
        if n == 0:
            result.scalars.return_value.all.return_value = [user]
        elif n == 1:
            result.scalars.return_value.all.return_value = [acct]
        elif n == 2:
            result.scalars.return_value.all.return_value = [prog]
        elif n == 3:
            result.scalars.return_value.all.return_value = []  # no existing assignments
        else:
            result.scalars.return_value.all.return_value = []
        return result

    db.execute = mock_execute

    headers = ["BDM Email *", "Account *", "Program *"]
    rows_data = [
        ["bdm@example.com", "Acme Corp", "Q1 Promo"],
    ]
    xlsx = _make_xlsx_bytes([headers] + rows_data)

    rows = await parse_and_validate_entities(db, xlsx, "assignments")
    assert len(rows) == 1
    assert rows[0].status == "ok"


@pytest.mark.asyncio
async def test_import_assignments_unknown_bdm():
    """Assignment with unknown BDM email should be an error."""
    acct = _make_account("Acme Corp")
    prog = _make_program("Q1 Promo", acct.id)

    db = AsyncMock()
    call_idx = [0]

    async def mock_execute(stmt, *args, **kwargs):
        n = call_idx[0]
        call_idx[0] += 1
        result = MagicMock()
        if n == 0:
            result.scalars.return_value.all.return_value = []  # no users
        elif n == 1:
            result.scalars.return_value.all.return_value = [acct]
        elif n == 2:
            result.scalars.return_value.all.return_value = [prog]
        elif n == 3:
            result.scalars.return_value.all.return_value = []
        else:
            result.scalars.return_value.all.return_value = []
        return result

    db.execute = mock_execute

    headers = ["BDM Email *", "Account *", "Program *"]
    rows_data = [
        ["unknown@example.com", "Acme Corp", "Q1 Promo"],
    ]
    xlsx = _make_xlsx_bytes([headers] + rows_data)

    rows = await parse_and_validate_entities(db, xlsx, "assignments")
    assert len(rows) == 1
    assert rows[0].status == "error"
    assert "User" in rows[0].error_msg
