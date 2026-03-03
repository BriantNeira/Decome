"""KPI calculation service — summaries, Excel export, AI diagnosis."""
from __future__ import annotations

import datetime
import io
import uuid
from collections import defaultdict

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import sqlalchemy as sa
from sqlalchemy import case, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.account import Account
from app.models.generated_message import GeneratedMessage
from app.models.program import Program
from app.models.reminder import Reminder
from app.models.reminder_type import ReminderType
from app.models.token_usage_log import TokenUsageLog
from app.models.user import User
from app.schemas.kpi import (
    KPIAccountSummary,
    KPIBDMSummary,
    KPIProgramSummary,
    KPISummary,
    KPITokenBDM,
    KPITokenSummary,
    KPITypeSummary,
)


# ---------------------------------------------------------------------------
# Main KPI aggregation
# ---------------------------------------------------------------------------

async def get_kpi_summary(
    db: AsyncSession,
    *,
    date_from: str,
    date_to: str,
    account_id: uuid.UUID | None = None,
    program_id: uuid.UUID | None = None,
    bdm_id: uuid.UUID | None = None,
) -> KPISummary:
    d_from = datetime.date.fromisoformat(date_from)
    d_to = datetime.date.fromisoformat(date_to)
    today = datetime.date.today()

    # ── base filter ────────────────────────────────────────────────────
    base_filters = [
        Reminder.start_date >= d_from,
        Reminder.start_date <= d_to,
    ]
    if account_id:
        base_filters.append(Reminder.account_id == account_id)
    if program_id:
        base_filters.append(Reminder.program_id == program_id)
    if bdm_id:
        base_filters.append(Reminder.user_id == bdm_id)

    # ── load all matching reminders (with relationships) ───────────────
    stmt = select(Reminder).where(*base_filters).order_by(Reminder.start_date)
    result = await db.execute(stmt)
    reminders = list(result.scalars().all())

    # ── scalar KPIs ────────────────────────────────────────────────────
    completed_on_time = 0
    completed_late = 0
    overdue_pending = 0
    total_open = 0
    total_completed = 0

    for r in reminders:
        if r.status == "completed":
            total_completed += 1
            # "On time" means completed_at exists and falls on or before end-of-day
            # of start_date
            if r.completed_at is not None:
                completed_date = r.completed_at.date() if hasattr(r.completed_at, "date") else r.completed_at
                if completed_date <= r.start_date:
                    completed_on_time += 1
                else:
                    completed_late += 1
            else:
                completed_late += 1
        elif r.status in ("open", "in_progress"):
            total_open += 1
            if r.start_date < today:
                overdue_pending += 1

    total_completions = completed_on_time + completed_late
    completion_rate = round(completed_on_time / total_completions * 100, 1) if total_completions else 0.0

    # ── by_type ────────────────────────────────────────────────────────
    type_map: dict[int | None, dict] = defaultdict(lambda: {"total": 0, "completed": 0, "overdue": 0, "name": "Untyped", "color": None})
    for r in reminders:
        key = r.type_id
        if r.reminder_type:
            type_map[key]["name"] = r.reminder_type.name
            type_map[key]["color"] = r.reminder_type.color
        type_map[key]["total"] += 1
        if r.status == "completed":
            type_map[key]["completed"] += 1
        if r.status in ("open", "in_progress") and r.start_date < today:
            type_map[key]["overdue"] += 1

    by_type = [
        KPITypeSummary(
            type_id=k,
            type_name=v["name"],
            type_color=v["color"],
            total=v["total"],
            completed=v["completed"],
            overdue=v["overdue"],
        )
        for k, v in type_map.items()
    ]

    # ── by_account ─────────────────────────────────────────────────────
    acct_map: dict[uuid.UUID, dict] = defaultdict(lambda: {"name": "", "total": 0, "completed": 0, "overdue": 0, "on_time": 0})
    for r in reminders:
        key = r.account_id
        acct_map[key]["name"] = r.account.name if r.account else str(key)
        acct_map[key]["total"] += 1
        if r.status == "completed":
            acct_map[key]["completed"] += 1
            if r.completed_at and r.completed_at.date() <= r.start_date:
                acct_map[key]["on_time"] += 1
        if r.status in ("open", "in_progress") and r.start_date < today:
            acct_map[key]["overdue"] += 1

    by_account = [
        KPIAccountSummary(
            account_id=str(k),
            account_name=v["name"],
            total=v["total"],
            completed=v["completed"],
            overdue=v["overdue"],
            on_time_pct=round(v["on_time"] / v["completed"] * 100, 1) if v["completed"] else 0.0,
        )
        for k, v in acct_map.items()
    ]

    # ── by_program ─────────────────────────────────────────────────────
    prog_map: dict[uuid.UUID | None, dict] = defaultdict(lambda: {"prog_name": "No Program", "acct_name": "", "total": 0, "completed": 0, "overdue": 0})
    for r in reminders:
        key = r.program_id
        if r.program:
            prog_map[key]["prog_name"] = r.program.name
            prog_map[key]["acct_name"] = r.program.account.name if r.program.account else ""
        else:
            prog_map[key]["acct_name"] = r.account.name if r.account else ""
        prog_map[key]["total"] += 1
        if r.status == "completed":
            prog_map[key]["completed"] += 1
        if r.status in ("open", "in_progress") and r.start_date < today:
            prog_map[key]["overdue"] += 1

    by_program = [
        KPIProgramSummary(
            program_id=str(k) if k else None,
            program_name=v["prog_name"],
            account_name=v["acct_name"],
            total=v["total"],
            completed=v["completed"],
            overdue=v["overdue"],
        )
        for k, v in prog_map.items()
    ]

    # ── by_bdm (including token + message counts) ──────────────────────
    bdm_map: dict[uuid.UUID, dict] = defaultdict(lambda: {"name": "", "email": "", "total": 0, "completed": 0, "overdue": 0})
    for r in reminders:
        key = r.user_id
        if r.user:
            bdm_map[key]["name"] = r.user.full_name
            bdm_map[key]["email"] = r.user.email
        bdm_map[key]["total"] += 1
        if r.status == "completed":
            bdm_map[key]["completed"] += 1
        if r.status in ("open", "in_progress") and r.start_date < today:
            bdm_map[key]["overdue"] += 1

    # Token usage per BDM in date range
    token_stmt = (
        select(
            TokenUsageLog.user_id,
            func.coalesce(func.sum(TokenUsageLog.tokens_used), 0).label("tokens"),
        )
        .where(
            func.cast(TokenUsageLog.generated_at, sa.Date) >= d_from,
            func.cast(TokenUsageLog.generated_at, sa.Date) <= d_to,
        )
        .group_by(TokenUsageLog.user_id)
    )
    token_result = await db.execute(token_stmt)
    token_rows = {row.user_id: int(row.tokens) for row in token_result}

    # Generated messages per BDM in date range
    msg_stmt = (
        select(
            GeneratedMessage.generated_by,
            func.count().label("cnt"),
        )
        .where(
            func.cast(GeneratedMessage.generated_at, sa.Date) >= d_from,
            func.cast(GeneratedMessage.generated_at, sa.Date) <= d_to,
        )
        .group_by(GeneratedMessage.generated_by)
    )
    msg_result = await db.execute(msg_stmt)
    msg_rows = {row.generated_by: int(row.cnt) for row in msg_result}

    by_bdm = [
        KPIBDMSummary(
            user_id=str(uid),
            user_name=v["name"],
            user_email=v["email"],
            total=v["total"],
            completed=v["completed"],
            overdue=v["overdue"],
            tokens_used=token_rows.get(uid, 0),
            messages_generated=msg_rows.get(uid, 0),
        )
        for uid, v in bdm_map.items()
    ]

    # ── token_summary ──────────────────────────────────────────────────
    total_tokens = sum(token_rows.values())
    token_by_bdm: list[KPITokenBDM] = []
    # Collect all user_ids that appear in token or message rows
    all_bdm_ids = set(token_rows.keys()) | set(msg_rows.keys())
    if all_bdm_ids:
        users_stmt = select(User).where(User.id.in_(all_bdm_ids))
        users_result = await db.execute(users_stmt)
        users_by_id = {u.id: u for u in users_result.scalars().all()}
    else:
        users_by_id = {}

    for uid in all_bdm_ids:
        u = users_by_id.get(uid)
        token_by_bdm.append(KPITokenBDM(
            name=u.full_name if u else str(uid),
            tokens=token_rows.get(uid, 0),
            messages=msg_rows.get(uid, 0),
        ))

    token_summary = KPITokenSummary(total_tokens=total_tokens, by_bdm=token_by_bdm)

    return KPISummary(
        date_from=date_from,
        date_to=date_to,
        completed_on_time=completed_on_time,
        completed_late=completed_late,
        completion_rate=completion_rate,
        overdue_pending=overdue_pending,
        total_open=total_open,
        total_completed=total_completed,
        by_type=by_type,
        by_account=by_account,
        by_program=by_program,
        by_bdm=by_bdm,
        token_summary=token_summary,
    )


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="9AAE2F", end_color="9AAE2F", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


def _write_header(ws, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _ALIGN_CENTER


async def export_calendar_excel(
    db: AsyncSession,
    *,
    date_from: str,
    date_to: str,
    account_id: uuid.UUID | None = None,
    program_id: uuid.UUID | None = None,
    bdm_id: uuid.UUID | None = None,
) -> bytes:
    summary = await get_kpi_summary(
        db, date_from=date_from, date_to=date_to,
        account_id=account_id, program_id=program_id, bdm_id=bdm_id,
    )

    d_from = datetime.date.fromisoformat(date_from)
    d_to = datetime.date.fromisoformat(date_to)

    # Fetch all reminders for the detail sheet
    base_filters = [
        Reminder.start_date >= d_from,
        Reminder.start_date <= d_to,
    ]
    if account_id:
        base_filters.append(Reminder.account_id == account_id)
    if program_id:
        base_filters.append(Reminder.program_id == program_id)
    if bdm_id:
        base_filters.append(Reminder.user_id == bdm_id)

    result = await db.execute(
        select(Reminder).where(*base_filters).order_by(Reminder.start_date)
    )
    reminders = list(result.scalars().all())

    wb = Workbook()

    # ── Sheet 1: Summary ───────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_header(ws_summary, ["Metric", "Value"])
    rows = [
        ("Date Range", f"{date_from} to {date_to}"),
        ("Completed On Time", summary.completed_on_time),
        ("Completed Late", summary.completed_late),
        ("Completion Rate (%)", summary.completion_rate),
        ("Overdue Pending", summary.overdue_pending),
        ("Total Open", summary.total_open),
        ("Total Completed", summary.total_completed),
        ("Total Tokens Used", summary.token_summary.total_tokens),
    ]
    for row_idx, (metric, value) in enumerate(rows, 2):
        ws_summary.cell(row=row_idx, column=1, value=metric)
        ws_summary.cell(row=row_idx, column=2, value=value)
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 30

    # ── Sheet 2: By Account (pivot with months) ───────────────────────
    ws_acct = wb.create_sheet("By Account")
    # Build month columns between d_from and d_to
    months = _month_range(d_from, d_to)
    month_labels = [m.strftime("%Y-%m") for m in months]
    _write_header(ws_acct, ["Account"] + month_labels + ["Total"])

    acct_month: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in reminders:
        acct_name = r.account.name if r.account else "Unknown"
        month_key = r.start_date.strftime("%Y-%m")
        acct_month[acct_name][month_key] += 1

    for row_idx, (acct_name, month_counts) in enumerate(sorted(acct_month.items()), 2):
        ws_acct.cell(row=row_idx, column=1, value=acct_name)
        total = 0
        for col_idx, ml in enumerate(month_labels, 2):
            cnt = month_counts.get(ml, 0)
            ws_acct.cell(row=row_idx, column=col_idx, value=cnt)
            total += cnt
        ws_acct.cell(row=row_idx, column=len(month_labels) + 2, value=total)
    ws_acct.column_dimensions["A"].width = 30

    # ── Sheet 3: By BDM (pivot with months) ───────────────────────────
    ws_bdm = wb.create_sheet("By BDM")
    _write_header(ws_bdm, ["BDM"] + month_labels + ["Total"])

    bdm_month: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in reminders:
        bdm_name = r.user.full_name if r.user else "Unknown"
        month_key = r.start_date.strftime("%Y-%m")
        bdm_month[bdm_name][month_key] += 1

    for row_idx, (bdm_name, month_counts) in enumerate(sorted(bdm_month.items()), 2):
        ws_bdm.cell(row=row_idx, column=1, value=bdm_name)
        total = 0
        for col_idx, ml in enumerate(month_labels, 2):
            cnt = month_counts.get(ml, 0)
            ws_bdm.cell(row=row_idx, column=col_idx, value=cnt)
            total += cnt
        ws_bdm.cell(row=row_idx, column=len(month_labels) + 2, value=total)
    ws_bdm.column_dimensions["A"].width = 30

    # ── Sheet 4: All Reminders ─────────────────────────────────────────
    ws_all = wb.create_sheet("All Reminders")
    detail_headers = ["Date", "Account", "Program", "Type", "Title", "Status", "BDM", "Completed At"]
    _write_header(ws_all, detail_headers)

    for row_idx, r in enumerate(reminders, 2):
        ws_all.cell(row=row_idx, column=1, value=str(r.start_date))
        ws_all.cell(row=row_idx, column=2, value=r.account.name if r.account else "")
        ws_all.cell(row=row_idx, column=3, value=r.program.name if r.program else "")
        ws_all.cell(row=row_idx, column=4, value=r.reminder_type.name if r.reminder_type else "")
        ws_all.cell(row=row_idx, column=5, value=r.title)
        ws_all.cell(row=row_idx, column=6, value=r.status)
        ws_all.cell(row=row_idx, column=7, value=r.user.full_name if r.user else "")
        ws_all.cell(row=row_idx, column=8, value=str(r.completed_at) if r.completed_at else "")

    for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws_all.column_dimensions[col_letter].width = 20

    # ── Save to bytes ──────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _month_range(d_from: datetime.date, d_to: datetime.date) -> list[datetime.date]:
    """Return a list of first-of-month dates covering the range."""
    months: list[datetime.date] = []
    current = d_from.replace(day=1)
    end = d_to.replace(day=1)
    while current <= end:
        months.append(current)
        # Advance to next month
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)
    return months


# ---------------------------------------------------------------------------
# AI Diagnosis
# ---------------------------------------------------------------------------

async def generate_ai_diagnosis(
    db: AsyncSession,
    *,
    user_id: uuid.UUID,
    date_from: str,
    date_to: str,
    account_id: uuid.UUID | None = None,
) -> tuple[str, int]:
    """Build KPI summary, send to LLM for diagnosis, log usage. Returns (text, tokens)."""
    from app.services import budget_service, llm_service

    # 1. Budget check
    await budget_service.check_budget(db, user_id)

    # 2. LLM config
    config = await llm_service.get_llm_config(db)
    if config is None or not config.is_active:
        raise HTTPException(status_code=422, detail="LLM is not configured or not active.")
    if not config.api_key:
        raise HTTPException(status_code=422, detail="LLM API key is not set.")

    # 3. Get KPI summary
    summary = await get_kpi_summary(
        db, date_from=date_from, date_to=date_to, account_id=account_id,
    )

    # 4. Build text for LLM
    lines = [
        f"Period: {summary.date_from} to {summary.date_to}",
        f"Completed on time: {summary.completed_on_time}",
        f"Completed late: {summary.completed_late}",
        f"Completion rate: {summary.completion_rate}%",
        f"Overdue pending: {summary.overdue_pending}",
        f"Total open: {summary.total_open}",
        f"Total completed: {summary.total_completed}",
        "",
        "By Type:",
    ]
    for t in summary.by_type:
        lines.append(f"  - {t.type_name}: {t.total} total, {t.completed} completed, {t.overdue} overdue")

    lines.append("")
    lines.append("By Account:")
    for a in summary.by_account:
        lines.append(f"  - {a.account_name}: {a.total} total, {a.completed} completed, {a.overdue} overdue, {a.on_time_pct}% on-time")

    lines.append("")
    lines.append("By BDM:")
    for b in summary.by_bdm:
        lines.append(
            f"  - {b.user_name}: {b.total} total, {b.completed} completed, "
            f"{b.overdue} overdue, {b.tokens_used} tokens, {b.messages_generated} messages"
        )

    lines.append("")
    lines.append(f"Token usage: {summary.token_summary.total_tokens} total")

    kpi_text = "\n".join(lines)

    # 5. Call LLM
    diagnosis_text, tokens_used = await llm_service.generate_diagnosis(config, kpi_text)

    # 6. Log usage
    await budget_service.log_usage(
        db,
        user_id=user_id,
        account_id=account_id,
        reminder_id=None,
        tokens=tokens_used,
        provider=config.provider,
        model=config.model,
    )

    return diagnosis_text, tokens_used
