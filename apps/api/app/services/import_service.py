"""Bulk import from Excel (.xlsx) — reminders and multi-entity."""
from __future__ import annotations

import io
import uuid
import datetime
from dataclasses import dataclass, field
from typing import Optional

from sqlalchemy import select, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.account import Account
from app.models.program import Program
from app.models.reminder import Reminder
from app.models.reminder_type import ReminderType
from app.models.contact import Contact
from app.models.assignment import Assignment
from app.models.user import User


EXPECTED_HEADERS = ["account", "program", "reminder_type", "title", "due_date", "notes", "recurrence"]
REQUIRED_COLUMNS = {"account", "program", "reminder_type", "title", "due_date"}
VALID_RECURRENCE = {"none", "daily", "weekly", "biweekly", "monthly", "yearly", ""}

ENTITY_COLUMNS = {
    "accounts": ["name", "code", "description"],
    "programs": ["name", "description", "account", "season"],
    "contacts": ["account", "first_name", "last_name", "email", "phone", "title", "is_decision_maker"],
    "assignments": ["bdm_email", "account", "program"],
}


@dataclass
class ImportRow:
    row_num: int
    account: str = ""
    program: str = ""
    reminder_type: str = ""
    title: str = ""
    due_date_raw: str = ""
    notes: str = ""
    recurrence: str = ""
    # resolved FKs
    account_id: Optional[uuid.UUID] = None
    program_id: Optional[uuid.UUID] = None
    type_id: Optional[int] = None
    due_date: Optional[datetime.date] = None
    # result
    status: str = "ok"
    error_msg: Optional[str] = None
    reminder_id: Optional[uuid.UUID] = None
    entity_name: Optional[str] = None


def _parse_date(raw: str) -> Optional[datetime.date]:
    """Try multiple date formats: YYYY-MM-DD, M/D/YYYY, D/M/YYYY, Excel serial."""
    if not raw:
        return None
    raw = str(raw).strip()

    # Try standard ISO
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y"):
        try:
            return datetime.datetime.strptime(raw, fmt).date()
        except ValueError:
            pass

    # Try Excel serial date (integer)
    try:
        serial = int(float(raw))
        # Excel serial 1 = 1900-01-01 (with Lotus 1-2-3 leap year bug)
        excel_epoch = datetime.date(1899, 12, 30)
        return excel_epoch + datetime.timedelta(days=serial)
    except (ValueError, OverflowError):
        pass

    return None


def generate_template_xlsx() -> bytes:
    """Generate a .xlsx template file with headers and example rows."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reminders"

    # Header style
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ["account", "program", "reminder_type", "title", "due_date", "notes", "recurrence"]
    header_labels = ["Account *", "Program *", "Reminder Type *", "Title *", "Due Date *", "Notes", "Recurrence"]
    col_widths = [25, 20, 20, 35, 15, 35, 15]

    for col_num, (header, label, width) in enumerate(zip(headers, header_labels, col_widths), 1):
        cell = ws.cell(row=1, column=col_num, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = width

    # Example rows
    example_rows = [
        ["Acme Corp", "Q1 Promotion", "Seasonal Order", "Spring Order Reminder", "2026-04-01", "Ask about new SKUs", "none"],
        ["Global Trade", "Annual Review", "Reorder", "Restock Follow-up", "2026-05-15", "", "monthly"],
    ]
    for row_data in example_rows:
        ws.append(row_data)

    # Notes sheet
    notes_ws = wb.create_sheet("Instructions")
    notes_ws.append(["Field", "Required?", "Notes"])
    notes_ws.append(["account", "Yes", "Exact account name (case-insensitive)"])
    notes_ws.append(["program", "Yes", "Exact program name (case-insensitive)"])
    notes_ws.append(["reminder_type", "Yes", "Exact reminder type name (case-insensitive)"])
    notes_ws.append(["title", "Yes", "Reminder title"])
    notes_ws.append(["due_date", "Yes", "Format: YYYY-MM-DD or M/D/YYYY"])
    notes_ws.append(["notes", "No", "Optional notes"])
    notes_ws.append(["recurrence", "No", "none, daily, weekly, biweekly, monthly, yearly"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def parse_and_validate(
    db: AsyncSession,
    file_bytes: bytes,
) -> list[ImportRow]:
    """Parse xlsx bytes, validate each row. Returns list of ImportRow (with status ok/error)."""
    import openpyxl

    rows: list[ImportRow] = []

    # Load workbook
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not open Excel file: {e}")

    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no sheets.")

    # Read header row
    sheet_rows = list(ws.iter_rows(values_only=True))
    if not sheet_rows:
        raise ValueError("Excel file is empty.")

    raw_headers = [str(h).strip().lower().replace("*", "").strip() if h else "" for h in sheet_rows[0]]
    if not raw_headers or raw_headers[0] == "":
        raise ValueError("First row must be headers.")

    # Build column index map
    col_map: dict[str, int] = {}
    for idx, h in enumerate(raw_headers):
        if h in EXPECTED_HEADERS:
            col_map[h] = idx

    missing = REQUIRED_COLUMNS - set(col_map.keys())
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

    if len(sheet_rows) < 2:
        return []

    # Pre-load lookups for O(1) per row
    accs_result = await db.execute(select(Account).where(Account.is_active == True))  # noqa: E712
    accounts_map: dict[str, Account] = {a.name.lower(): a for a in accs_result.scalars().all()}

    progs_result = await db.execute(select(Program).where(Program.is_active == True))  # noqa: E712
    programs_by_account: dict[tuple[uuid.UUID, str], Program] = {}
    all_programs: dict[str, Program] = {}
    for p in progs_result.scalars().all():
        all_programs[p.name.lower()] = p
        programs_by_account[(p.account_id, p.name.lower())] = p

    types_result = await db.execute(select(ReminderType).where(ReminderType.is_active == True))  # noqa: E712
    types_map: dict[str, ReminderType] = {t.name.lower(): t for t in types_result.scalars().all()}

    def get_cell(row_data: tuple, col_key: str) -> str:
        idx = col_map.get(col_key)
        if idx is None or idx >= len(row_data):
            return ""
        v = row_data[idx]
        return str(v).strip() if v is not None else ""

    for row_num, row_data in enumerate(sheet_rows[1:], start=2):
        # Skip entirely blank rows
        if all(v is None or str(v).strip() == "" for v in row_data):
            continue

        ir = ImportRow(row_num=row_num)
        ir.account = get_cell(row_data, "account")
        ir.program = get_cell(row_data, "program")
        ir.reminder_type = get_cell(row_data, "reminder_type")
        ir.title = get_cell(row_data, "title")
        ir.due_date_raw = get_cell(row_data, "due_date")
        ir.notes = get_cell(row_data, "notes")
        ir.recurrence = get_cell(row_data, "recurrence").lower() or "none"

        errors = []

        # Validate required fields
        for field_name in ["account", "program", "reminder_type", "title", "due_date_raw"]:
            if not getattr(ir, field_name):
                errors.append(f"'{field_name.replace('_raw', '')}' is required")

        # Resolve account
        if ir.account:
            acc = accounts_map.get(ir.account.lower())
            if acc is None:
                errors.append(f"Account '{ir.account}' not found")
            else:
                ir.account_id = acc.id

        # Resolve program (scoped to account if found)
        if ir.program and ir.account_id:
            prog = programs_by_account.get((ir.account_id, ir.program.lower()))
            if prog is None:
                # Try global (program exists but may belong to different account)
                prog_global = all_programs.get(ir.program.lower())
                if prog_global is None:
                    errors.append(f"Program '{ir.program}' not found for account '{ir.account}'")
                else:
                    ir.program_id = prog_global.id
            else:
                ir.program_id = prog.id

        # Resolve reminder type
        if ir.reminder_type:
            rt = types_map.get(ir.reminder_type.lower())
            if rt is None:
                errors.append(f"Reminder type '{ir.reminder_type}' not found")
            else:
                ir.type_id = rt.id

        # Parse date
        if ir.due_date_raw:
            parsed_date = _parse_date(ir.due_date_raw)
            if parsed_date is None:
                errors.append(f"Invalid due_date '{ir.due_date_raw}' — use YYYY-MM-DD")
            else:
                ir.due_date = parsed_date

        # Validate recurrence
        if ir.recurrence not in VALID_RECURRENCE:
            errors.append(f"Invalid recurrence '{ir.recurrence}' — use: none, daily, weekly, biweekly, monthly, yearly")

        if errors:
            ir.status = "error"
            ir.error_msg = "; ".join(errors)
        else:
            ir.status = "ok"

        rows.append(ir)

    return rows


async def create_from_rows(
    db: AsyncSession,
    rows: list[ImportRow],
    created_by_id: uuid.UUID,
) -> list[ImportRow]:
    """Create reminders for all rows with status='ok'. Updates reminder_id in place."""
    for ir in rows:
        if ir.status != "ok":
            continue
        recurrence_rule = ir.recurrence.upper() if ir.recurrence and ir.recurrence != "none" else None
        reminder = Reminder(
            user_id=created_by_id,
            account_id=ir.account_id,
            program_id=ir.program_id,
            type_id=ir.type_id,
            title=ir.title,
            notes=ir.notes or None,
            start_date=ir.due_date,
            recurrence_rule=recurrence_rule,
            status="open",
        )
        db.add(reminder)
        await db.flush()  # get the ID
        ir.reminder_id = reminder.id

    await db.commit()
    return rows


# ---------------------------------------------------------------------------
# Multi-entity import helpers
# ---------------------------------------------------------------------------

_ENTITY_EXAMPLE_ROWS = {
    "accounts": [
        ["Acme Corp", "ACME-01", "Major retail partner"],
        ["Global Trade", "GT-100", "International distributor"],
    ],
    "programs": [
        ["Q1 Promotion", "First quarter promotion program", "Acme Corp", "Spring 2026"],
        ["Annual Review", "Yearly performance review", "", "Fall 2026"],
    ],
    "contacts": [
        ["Acme Corp", "Jane", "Doe", "jane@acme.com", "555-1234", "Buyer", "true"],
        ["Global Trade", "John", "Smith", "john@globaltrade.com", "555-5678", "Manager", "false"],
    ],
    "assignments": [
        ["bdm@example.com", "Acme Corp", "Q1 Promotion"],
        ["bdm@example.com", "Global Trade", "Annual Review"],
    ],
}

_ENTITY_HEADER_LABELS = {
    "accounts": ["Name *", "Code", "Description"],
    "programs": ["Name *", "Description", "Account", "Season"],
    "contacts": ["Account *", "First Name *", "Last Name *", "Email", "Phone", "Title", "Decision Maker"],
    "assignments": ["BDM Email *", "Account *", "Program *"],
}

_ENTITY_COL_WIDTHS = {
    "accounts": [30, 15, 40],
    "programs": [30, 40, 25, 15],
    "contacts": [25, 20, 20, 30, 15, 15, 18],
    "assignments": [30, 25, 25],
}

_ENTITY_INSTRUCTIONS = {
    "accounts": [
        ["Field", "Required?", "Notes"],
        ["name", "Yes", "Account name (must be unique)"],
        ["code", "No", "Account code (must be unique if provided)"],
        ["description", "No", "Optional description"],
    ],
    "programs": [
        ["Field", "Required?", "Notes"],
        ["name", "Yes", "Program name (must be unique)"],
        ["description", "No", "Optional description"],
        ["account", "No", "Account name to link (case-insensitive)"],
        ["season", "No", "Season label (e.g. Spring 2026)"],
    ],
    "contacts": [
        ["Field", "Required?", "Notes"],
        ["account", "Yes", "Account name (case-insensitive, must exist)"],
        ["first_name", "Yes", "Contact first name"],
        ["last_name", "Yes", "Contact last name"],
        ["email", "No", "Email address"],
        ["phone", "No", "Phone number"],
        ["title", "No", "Job title / role"],
        ["is_decision_maker", "No", "true or false (default false)"],
    ],
    "assignments": [
        ["Field", "Required?", "Notes"],
        ["bdm_email", "Yes", "BDM user email (must be existing user)"],
        ["account", "Yes", "Account name (case-insensitive)"],
        ["program", "Yes", "Program name (case-insensitive)"],
    ],
}


def generate_entity_template_xlsx(entity_type: str) -> bytes:
    """Generate a .xlsx template file for the given entity type."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    if entity_type not in ENTITY_COLUMNS:
        raise ValueError(f"Unknown entity type: {entity_type}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = entity_type.capitalize()

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ENTITY_COLUMNS[entity_type]
    labels = _ENTITY_HEADER_LABELS[entity_type]
    widths = _ENTITY_COL_WIDTHS[entity_type]

    for col_num, (header, label, width) in enumerate(zip(headers, labels, widths), 1):
        cell = ws.cell(row=1, column=col_num, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = width

    for row_data in _ENTITY_EXAMPLE_ROWS[entity_type]:
        ws.append(row_data)

    # Instructions sheet
    notes_ws = wb.create_sheet("Instructions")
    for row_data in _ENTITY_INSTRUCTIONS[entity_type]:
        notes_ws.append(row_data)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def parse_and_validate_entities(
    db: AsyncSession,
    file_bytes: bytes,
    entity_type: str,
) -> list[ImportRow]:
    """Parse xlsx bytes for a given entity type, validate each row."""
    import openpyxl

    if entity_type not in ENTITY_COLUMNS:
        raise ValueError(f"Unknown entity type: {entity_type}")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not open Excel file: {e}")

    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no sheets.")

    sheet_rows = list(ws.iter_rows(values_only=True))
    if not sheet_rows:
        raise ValueError("Excel file is empty.")

    expected = ENTITY_COLUMNS[entity_type]
    raw_headers = [
        str(h).strip().lower().replace("*", "").strip() if h else ""
        for h in sheet_rows[0]
    ]

    # Build column index map
    col_map: dict[str, int] = {}
    for idx, h in enumerate(raw_headers):
        # Normalise header: "decision maker" -> "is_decision_maker", "bdm email" -> "bdm_email"
        normalised = h.replace(" ", "_")
        if normalised in expected:
            col_map[normalised] = idx
        elif h in expected:
            col_map[h] = idx

    missing = set(expected) - set(col_map.keys())
    # Only require columns that have a '*' label
    required_set = {
        expected[i]
        for i, lbl in enumerate(_ENTITY_HEADER_LABELS[entity_type])
        if "*" in lbl
    }
    missing_required = required_set - set(col_map.keys())
    if missing_required:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing_required))}")

    if len(sheet_rows) < 2:
        return []

    def get_cell(row_data: tuple, col_key: str) -> str:
        idx = col_map.get(col_key)
        if idx is None or idx >= len(row_data):
            return ""
        v = row_data[idx]
        return str(v).strip() if v is not None else ""

    # Dispatch to entity-specific validator
    if entity_type == "accounts":
        return await _validate_accounts(db, sheet_rows[1:], get_cell)
    elif entity_type == "programs":
        return await _validate_programs(db, sheet_rows[1:], get_cell)
    elif entity_type == "contacts":
        return await _validate_contacts(db, sheet_rows[1:], get_cell)
    elif entity_type == "assignments":
        return await _validate_assignments(db, sheet_rows[1:], get_cell)
    else:
        raise ValueError(f"Unsupported entity type: {entity_type}")


async def _validate_accounts(db, data_rows, get_cell) -> list[ImportRow]:
    # Pre-load existing accounts for duplicate checking
    result = await db.execute(select(Account))
    existing_names = {a.name.lower() for a in result.scalars().all()}

    rows: list[ImportRow] = []
    seen_names: set[str] = set()  # Track names within this import batch

    for row_num, row_data in enumerate(data_rows, start=2):
        if all(v is None or str(v).strip() == "" for v in row_data):
            continue

        ir = ImportRow(row_num=row_num)
        name = get_cell(row_data, "name")
        code = get_cell(row_data, "code")
        description = get_cell(row_data, "description")

        ir.account = name  # store name in the account field for display
        ir.title = code  # reuse title field for code display
        ir.notes = description

        errors = []
        if not name:
            errors.append("'name' is required")

        if name:
            name_lower = name.lower()
            if name_lower in existing_names:
                ir.status = "skipped"
                ir.error_msg = f"Account '{name}' already exists"
                rows.append(ir)
                continue
            if name_lower in seen_names:
                ir.status = "skipped"
                ir.error_msg = f"Duplicate account '{name}' in this file"
                rows.append(ir)
                continue
            seen_names.add(name_lower)

        if errors:
            ir.status = "error"
            ir.error_msg = "; ".join(errors)
        else:
            ir.status = "ok"

        rows.append(ir)

    return rows


async def _validate_programs(db, data_rows, get_cell) -> list[ImportRow]:
    # Pre-load existing programs
    prog_result = await db.execute(select(Program))
    existing_names = {p.name.lower() for p in prog_result.scalars().all()}

    # Pre-load accounts for resolving optional account link
    acc_result = await db.execute(select(Account).where(Account.is_active == True))  # noqa: E712
    accounts_map: dict[str, Account] = {a.name.lower(): a for a in acc_result.scalars().all()}

    rows: list[ImportRow] = []
    seen_names: set[str] = set()

    for row_num, row_data in enumerate(data_rows, start=2):
        if all(v is None or str(v).strip() == "" for v in row_data):
            continue

        ir = ImportRow(row_num=row_num)
        name = get_cell(row_data, "name")
        description = get_cell(row_data, "description")
        account_name = get_cell(row_data, "account")
        season = get_cell(row_data, "season")

        ir.title = name  # reuse title for display
        ir.account = account_name
        ir.notes = description
        ir.recurrence = season  # reuse recurrence field for season display

        errors = []
        if not name:
            errors.append("'name' is required")

        if name:
            name_lower = name.lower()
            if name_lower in existing_names or name_lower in seen_names:
                ir.status = "skipped"
                ir.error_msg = f"Program '{name}' already exists"
                rows.append(ir)
                continue
            seen_names.add(name_lower)

        # Resolve optional account
        if account_name:
            acc = accounts_map.get(account_name.lower())
            if acc is None:
                errors.append(f"Account '{account_name}' not found")
            else:
                ir.account_id = acc.id

        if errors:
            ir.status = "error"
            ir.error_msg = "; ".join(errors)
        else:
            ir.status = "ok"

        rows.append(ir)

    return rows


async def _validate_contacts(db, data_rows, get_cell) -> list[ImportRow]:
    # Pre-load accounts
    acc_result = await db.execute(select(Account).where(Account.is_active == True))  # noqa: E712
    accounts_map: dict[str, Account] = {a.name.lower(): a for a in acc_result.scalars().all()}

    rows: list[ImportRow] = []

    for row_num, row_data in enumerate(data_rows, start=2):
        if all(v is None or str(v).strip() == "" for v in row_data):
            continue

        ir = ImportRow(row_num=row_num)
        account_name = get_cell(row_data, "account")
        first_name = get_cell(row_data, "first_name")
        last_name = get_cell(row_data, "last_name")
        email = get_cell(row_data, "email")
        phone = get_cell(row_data, "phone")
        title = get_cell(row_data, "title")
        is_dm = get_cell(row_data, "is_decision_maker").lower()

        ir.account = account_name
        ir.title = f"{first_name} {last_name}".strip()
        ir.notes = email
        ir.program = phone  # reuse program field for phone display
        ir.reminder_type = title  # reuse reminder_type for job title display
        ir.recurrence = is_dm  # reuse recurrence for is_decision_maker display

        errors = []
        if not account_name:
            errors.append("'account' is required")
        if not first_name:
            errors.append("'first_name' is required")
        if not last_name:
            errors.append("'last_name' is required")

        # Resolve account (required)
        if account_name:
            acc = accounts_map.get(account_name.lower())
            if acc is None:
                errors.append(f"Account '{account_name}' not found")
            else:
                ir.account_id = acc.id

        if errors:
            ir.status = "error"
            ir.error_msg = "; ".join(errors)
        else:
            ir.status = "ok"

        rows.append(ir)

    return rows


async def _validate_assignments(db, data_rows, get_cell) -> list[ImportRow]:
    # Pre-load users (BDMs)
    users_result = await db.execute(select(User).where(User.is_active == True))  # noqa: E712
    users_map: dict[str, User] = {u.email.lower(): u for u in users_result.scalars().all()}

    # Pre-load accounts
    acc_result = await db.execute(select(Account).where(Account.is_active == True))  # noqa: E712
    accounts_map: dict[str, Account] = {a.name.lower(): a for a in acc_result.scalars().all()}

    # Pre-load programs
    prog_result = await db.execute(select(Program).where(Program.is_active == True))  # noqa: E712
    programs_map: dict[str, Program] = {p.name.lower(): p for p in prog_result.scalars().all()}

    # Pre-load existing assignments for duplicate detection
    assign_result = await db.execute(select(Assignment).where(Assignment.is_active == True))  # noqa: E712
    existing_assignments: set[tuple[uuid.UUID, uuid.UUID, uuid.UUID]] = set()
    for a in assign_result.scalars().all():
        existing_assignments.add((a.user_id, a.account_id, a.program_id))

    rows: list[ImportRow] = []

    for row_num, row_data in enumerate(data_rows, start=2):
        if all(v is None or str(v).strip() == "" for v in row_data):
            continue

        ir = ImportRow(row_num=row_num)
        bdm_email = get_cell(row_data, "bdm_email")
        account_name = get_cell(row_data, "account")
        program_name = get_cell(row_data, "program")

        ir.reminder_type = bdm_email  # reuse for display
        ir.account = account_name
        ir.program = program_name

        errors = []
        user_id = None

        if not bdm_email:
            errors.append("'bdm_email' is required")
        if not account_name:
            errors.append("'account' is required")
        if not program_name:
            errors.append("'program' is required")

        # Resolve BDM
        if bdm_email:
            user = users_map.get(bdm_email.lower())
            if user is None:
                errors.append(f"User '{bdm_email}' not found")
            else:
                user_id = user.id

        # Resolve account
        if account_name:
            acc = accounts_map.get(account_name.lower())
            if acc is None:
                errors.append(f"Account '{account_name}' not found")
            else:
                ir.account_id = acc.id

        # Resolve program
        if program_name:
            prog = programs_map.get(program_name.lower())
            if prog is None:
                errors.append(f"Program '{program_name}' not found")
            else:
                ir.program_id = prog.id

        # Check for duplicate assignment
        if user_id and ir.account_id and ir.program_id:
            if (user_id, ir.account_id, ir.program_id) in existing_assignments:
                ir.status = "skipped"
                ir.error_msg = "Assignment already exists"
                rows.append(ir)
                continue
            # Store user_id in type_id (int field won't work for uuid, so store in title)
            ir.title = str(user_id)

        if errors:
            ir.status = "error"
            ir.error_msg = "; ".join(errors)
        else:
            ir.status = "ok"
            # Store user_id in title for create step (if not already set above)
            if user_id and not ir.title:
                ir.title = str(user_id)

        rows.append(ir)

    return rows


async def create_entities_from_rows(
    db: AsyncSession,
    rows: list[ImportRow],
    entity_type: str,
    created_by_id: uuid.UUID,
) -> list[ImportRow]:
    """Create entities for all rows with status='ok'. Updates entity_name in place."""
    if entity_type == "accounts":
        return await _create_accounts(db, rows)
    elif entity_type == "programs":
        return await _create_programs(db, rows)
    elif entity_type == "contacts":
        return await _create_contacts(db, rows)
    elif entity_type == "assignments":
        return await _create_assignments(db, rows)
    else:
        raise ValueError(f"Unsupported entity type: {entity_type}")


async def _create_accounts(db: AsyncSession, rows: list[ImportRow]) -> list[ImportRow]:
    for ir in rows:
        if ir.status != "ok":
            continue
        account = Account(
            name=ir.account,
            code=ir.title if ir.title else None,  # code stored in title field
            description=ir.notes if ir.notes else None,
        )
        db.add(account)
        await db.flush()
        ir.entity_name = account.name
        ir.reminder_id = account.id  # reuse reminder_id for the created entity id
    await db.commit()
    return rows


async def _create_programs(db: AsyncSession, rows: list[ImportRow]) -> list[ImportRow]:
    for ir in rows:
        if ir.status != "ok":
            continue
        program = Program(
            name=ir.title,  # program name stored in title field
            description=ir.notes if ir.notes else None,
            account_id=ir.account_id,
            season=ir.recurrence if ir.recurrence else None,  # season stored in recurrence
        )
        db.add(program)
        await db.flush()
        ir.entity_name = program.name
        ir.reminder_id = program.id
    await db.commit()
    return rows


async def _create_contacts(db: AsyncSession, rows: list[ImportRow]) -> list[ImportRow]:
    for ir in rows:
        if ir.status != "ok":
            continue
        # Parse name from title field ("first last")
        parts = ir.title.split(" ", 1) if ir.title else ["", ""]
        first_name = parts[0]
        last_name = parts[1] if len(parts) > 1 else ""
        is_dm = ir.recurrence.lower() in ("true", "yes", "1") if ir.recurrence else False
        contact = Contact(
            account_id=ir.account_id,
            first_name=first_name,
            last_name=last_name,
            email=ir.notes if ir.notes else None,  # email stored in notes
            phone=ir.program if ir.program else None,  # phone stored in program
            title=ir.reminder_type if ir.reminder_type else None,  # job title stored in reminder_type
            is_decision_maker=is_dm,
        )
        db.add(contact)
        await db.flush()
        ir.entity_name = f"{contact.first_name} {contact.last_name}"
        ir.reminder_id = contact.id
    await db.commit()
    return rows


async def _create_assignments(db: AsyncSession, rows: list[ImportRow]) -> list[ImportRow]:
    for ir in rows:
        if ir.status != "ok":
            continue
        user_id = uuid.UUID(ir.title)  # stored during validation
        assignment = Assignment(
            user_id=user_id,
            account_id=ir.account_id,
            program_id=ir.program_id,
        )
        db.add(assignment)
        await db.flush()
        ir.entity_name = f"{ir.account} / {ir.program}"
        ir.reminder_id = assignment.id
    await db.commit()
    return rows
